import csv
import subprocess
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.data.config import (
    LEDGER_ACCOUNT_LIST,
    LEDGER_GSEC_ACCOUNT_LIST,
    LEDGER_ME_MAIN,
    LEDGER_MOM_MAIN,
    LEDGER_PAPA_MAIN,
    PORTFOLIO_EXPECTED_BALANCES,
    PORTFOLIO_RECON_REPORT,
)

EXPECTED_BALANCES: Dict[str, Dict[str, float]] = {}
with open(PORTFOLIO_EXPECTED_BALANCES, "r", encoding="utf-8") as f:
    filtered_lines = (
        line for line in f if line.strip() and not line.lstrip().startswith("#")
    )
    reader = csv.DictReader(filtered_lines)
    for row in reader:
        account = row["Account"].strip()
        period = row["Period"].strip()
        # print(f"Expected value for account: {account} and period: {period} is: {row["Expected"]}")
        expected = float(row["Expected"])
        EXPECTED_BALANCES.setdefault(account, {})[period] = expected


def read_account_list(account_files: List[Path]) -> List[str]:
    """
    Read multiple ledger account list files and return a sorted unique account list.
    """
    accounts: set[str] = set()

    for account_file in account_files:
        with open(account_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith(("#", ";")):
                    continue
                if line.startswith("account"):
                    line = line.replace("account", "").strip()
                accounts.add(line)

    return sorted(accounts)


def run_ledger_for_year_currency(
    currency: str,
    ledger_files: list,
    begin: str | None,
    end: str,
) -> Dict[str, float]:
    """
    Run one ledger balance command for ALL accounts for a given year & currency.
    Parses the default ledger output (no --format needed).
    """

    cmd = [
        "ledger",
        "balance",
        "-X",
        currency,
        "--strict",
        "-e",
        end,
    ]

    if begin:
        cmd.extend(["-b", begin])

    for lf in ledger_files:
        cmd.extend(["-f", str(lf)])

    # Debug print
    print(
        f"Running ledger balance for currency={currency} "
        f"from begin={begin or 'BEGINNING'} "
        f"to end={end}"
    )

    # Run ledger
    result = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        shell=False,
    )

    if result.returncode != 0:
        raise RuntimeError(f"Ledger error: {result.stderr.strip()}")

    balances = {}
    stack: List[str] = []
    for line in result.stdout.strip().splitlines():
        line = line.strip()
        if not line.strip():
            continue
        if line.strip().startswith("-") or line.strip() == "0":
            continue

        arr = line.split(" ")

        # filter out empty strings for clarity
        arr_filtered = [x for x in arr if x != ""]

        if len(arr_filtered) < 2:
            continue

        currency = arr_filtered[0]
        amount = float(arr_filtered[1].replace(",", ""))
        account_name = arr_filtered[-1]

        # number of empty strings
        level = len(arr) - 2 - (len(arr_filtered) - 1)
        # there are two spaces per level in ledger-cli output
        level = level // 2
        # print(f"{level}: {arr}")

        if level < len(stack):
            stack = stack[:level]  # truncate stack if needed

        if level == len(stack):
            # append new level
            stack.append(account_name)
        else:
            # replace existing level
            stack[level] = account_name

        full_account = ":".join(stack)
        # print(f"{full_account}: {amount}")
        balances[full_account] = amount

    return balances


def is_cumulative_account(account: str) -> bool:
    return account.startswith(("Assets", "Liabilities"))


def generate_reconciled_xlsx(
    account_list_file: List[Path],
    output_xlsx: Path,
    years: List[int],
    currencies: List[str],
    ledger_files: List[Path],
):
    """
    Generate CSV with accounts as rows, years/currencies as columns.
    Runs only one ledger CLI call per year per currency.
    """
    accounts = read_account_list(account_list_file)

    # all_balances[account][year-currency] = amount
    all_balances: Dict[str, Dict[str, float]] = {acct: {} for acct in accounts}

    for currency in currencies:
        for year in years:
            end = f"{year + 1}-01-01"

            cumulative_balances = run_ledger_for_year_currency(
                currency=currency,
                ledger_files=ledger_files,
                begin=None,
                end=end,
            )
            yearly_balances = run_ledger_for_year_currency(
                currency=currency,
                ledger_files=ledger_files,
                begin=f"{year}-01-01",
                end=end,
            )
            for account in accounts:
                key = f"{year}-{currency}"
                if is_cumulative_account(account):
                    balance = cumulative_balances.get(account, 0.0)
                else:
                    balance = yearly_balances.get(account, 0.0)
                all_balances[account][key] = balance

    # Prepare CSV headers
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    GREEN = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
        end_color="C6EFCE",
    )
    RED = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE",
    )
    WHITE = PatternFill(
        fill_type="solid",
        start_color="FFFFFF",
        end_color="FFFFFF",
    )
    HEADER_BLUE = PatternFill(
        fill_type="solid",
        start_color="4285F4",
        end_color="4285F4",
    )
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    ws.cell(row=1, column=1, value="Accounts")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=1, column=1).fill = HEADER_BLUE
    ws.cell(row=1, column=1).font = header_font

    col = 2
    for year in years:
        for currency in currencies:
            cell = ws.cell(row=1, column=col, value=f"{year}-{currency}")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = HEADER_BLUE
            cell.font = header_font
            col += 1
    # End Headers

    row = 2
    for account in accounts:
        ws.cell(row=row, column=1, value=account)

        col = 2
        for year in years:
            color = WHITE
            prev_col = col
            # check if expected value is defined in any of the currencies
            for currency in currencies:
                key = f"{year}-{currency}"
                actual = all_balances[account].get(key, 0.0)
                expected = EXPECTED_BALANCES.get(account, {}).get(key)

                cell = ws.cell(row=row, column=col, value=actual)
                cell.alignment = Alignment(horizontal="right")

                if expected is None:
                    if color != GREEN:
                        color = WHITE
                elif abs(actual - expected) < 0.01:
                    color = GREEN
                else:
                    if color != GREEN:
                        color = RED
                    print(
                        f"Mismatch → {account} {key}: expected={expected}, actual={actual}"
                    )
                col += 1

            # apply colors to all currencies in given year
            for c in range(prev_col, col):
                ws.cell(row=row, column=c).fill = color

        row += 1

    wb.save(output_xlsx)


if __name__ == "__main__":
    # -------------------------------
    # Config using project constants
    # -------------------------------
    ledger_files = [
        LEDGER_ME_MAIN,
        LEDGER_MOM_MAIN,
        LEDGER_PAPA_MAIN,
    ]

    years = list(range(2020, 2027))
    currencies = ["INR", "USD"]

    generate_reconciled_xlsx(
        account_list_file=[
            LEDGER_ACCOUNT_LIST,
            LEDGER_GSEC_ACCOUNT_LIST,
        ],
        output_xlsx=PORTFOLIO_RECON_REPORT,
        years=years,
        currencies=currencies,
        ledger_files=ledger_files,
    )

    print("✅ Reconciled XLSX generated successfully")
